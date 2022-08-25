# # -*-encoding:gbk-*-
import openpyxl
mianshi_list = [['name', 'goalPosition', 'changePosition', 'interviewTime', 'workStartTime',
                 'responses', 'performance', 'admissions']]  # 按行存放Excel表中数据
wb = openpyxl.load_workbook('./interview.xlsx')
ws = wb['面试情况']
maxrows = ws.max_row  # 获取最大行
lastflag = False
for i in range(maxrows-1):
    data = []
    for j, each in enumerate(ws.iter_cols(min_row=2, min_col=2, max_col=9)):
        if each[i].value == '录取情况\n（写出一篇后录用）':
            data.append('录取情况')
            continue
        if all([j == 1, each[i].value is None]):
            lastflag = True
            break
        else:
            data.append(each[i].value)
    if lastflag:
        break
    else:
        mianshi_list.append(data)
print(mianshi_list)

hao_list = [['name', 'account', 'password', 'cardNum', 'bank', 'idCard', 'phone']]
ws1 = wb['账号开通']
maxrows = ws1.max_row  # 获取最大行
lastflag = False
for i in range(maxrows-1):
    data = []
    for j, each in enumerate(ws1.iter_cols(min_row=2, min_col=2, max_col=4)):
        if all([j == 0, each[i].value is None]):
            lastflag = True
            break
        else:
            v = each[i].value
            if v is None:
                data.append(v)
                data.append(v)
            elif j == 1:
                account = v.split('\n')[0]
                password = v.split('\n')[1]
                data.append(account)
                data.append(password)
            elif j == 2:
                cardId = v.split('\n')[0].replace('卡号：', '')
                bank = v.split('\n')[1].replace('开户行：', '')
                idCard = v.split('\n')[2].replace('身份证：', '')
                phone = v.split('\n')[3].replace('手机号：', '')
                data.append(cardId)
                data.append(bank)
                data.append(idCard)
                data.append(phone)
            else:
                data.append(each[i].value)
    if lastflag:
        break
    else:
        hao_list.append(data)
print(hao_list)
wb.close()
