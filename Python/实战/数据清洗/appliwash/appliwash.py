# coding: utf-8
import os
import openpyxl

xlsx_path = './appli.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['简表']
appli_list = []
maxrows = ws.max_row  # 获取最大行
lastflag = False
for i in range(maxrows - 1):
    data = []
    for j, each in enumerate(ws.iter_cols(min_row=3, min_col=1, max_col=12)):
        d = each[i].value
        if all([j == 1, each[i].value is None]):
            lastflag = True
            break
            pass
        else:
            if d is None:
                d = ''
            data.append(d)
    if lastflag:
        break
    else:
        appli_list.append(data)
print(appli_list)