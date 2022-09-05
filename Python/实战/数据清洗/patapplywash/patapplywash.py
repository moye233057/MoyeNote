# coding: utf-8
import os
import openpyxl

xlsx_path = './patApply.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Sheet1']
first_appli_list = []
lastflag = False
for i in range(1, 15):
    data = []
    for j, each in enumerate(ws.iter_cols(min_row=2, min_col=1, max_col=32)):
        d = each[i].value
        if all([j == 9, each[i].value is None]):
            lastflag = True
            break
        else:
            if d is None:
                d = ''
            data.append(d)
    if lastflag:
        break
    else:
        first_appli_list.append(data)
print(first_appli_list)

multi_appli_list = []
lastflag = False
for i in range(1, 4):
    data = []
    for j, each in enumerate(ws.iter_cols(min_row=18, min_col=1, max_col=32)):
        d = each[i].value
        if all([j == 9, each[i].value is None]):
            lastflag = True
            break
        else:
            if d is None:
                d = ''
            data.append(d)
    if lastflag:
        break
    else:
        multi_appli_list.append(data)
print(multi_appli_list)

recover_appli_list = []
lastflag = False
for i in range(1, 3):
    data = []
    for j, each in enumerate(ws.iter_cols(min_row=24, min_col=1, max_col=32)):
        d = each[i].value
        if d is None:
            d = ''
        data.append(d)
    if lastflag:
        break
    else:
        recover_appli_list.append(data)
print(recover_appli_list)